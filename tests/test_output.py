"""Tests for Phase 1 output workbook/report compatibility."""

from __future__ import annotations

from datetime import datetime, timedelta, timezone

import openpyxl
import pandas as pd
from docx import Document

from src.config import ScraperConfig
from src.main import _compute_run_summary
from src.output.excel_writer import generate_excel
from src.output.word_report import generate_word_report


def test_excel_includes_phase1_summary_sheets_and_finanzierung_alias(tmp_path):
    path = tmp_path / "dashboard.xlsx"
    vendors = pd.DataFrame(
        [{"Händler ID": "C0000001", "Händlername": "Demo Händler", "Mobile.de_Links": "https://home.mobile.de/DEMO"}]
    )
    cars = pd.DataFrame(
        [
            {
                "Händler ID": "C0000001",
                "Markes": "Volkswagen",
                "Models": "Golf",
                "Financing": "100 € mtl.",
                "Vehicle_URL": "https://suchen.mobile.de/fahrzeuge/details.html?id=1",
            }
        ]
    )
    dashboard = {
        "vendor_summary": pd.DataFrame(),
        "manufacturer_summary": pd.DataFrame(),
        "category_summary": pd.DataFrame(),
        "best_deals": pd.DataFrame(),
        "worst_deals": pd.DataFrame(),
        "efficient_vehicles": pd.DataFrame(),
    }

    generate_excel(
        path,
        vendors,
        cars,
        cars,
        dashboard,
        run_summary={"run_id": "test-run", "vendors": 1, "vehicles_raw": 1},
        vendor_coverage=pd.DataFrame(
            [{"field": "Händlername", "non_empty_count": 1, "total_count": 1, "coverage_pct": 100.0}]
        ),
        vehicle_coverage=pd.DataFrame(
            [{"field": "Financing", "non_empty_count": 1, "total_count": 1, "coverage_pct": 100.0}]
        ),
        errors=[{"type": "vehicle", "url": "https://example.test", "error": "boom"}],
    )

    workbook = openpyxl.load_workbook(path)
    assert {
        "Vendors",
        "Vehicles",
        "Run_Summary",
        "Data_Coverage",
        "Requirements_Compliance",
        "Errors",
        "Classification_Summary",
    }.issubset(set(workbook.sheetnames))
    raw_headers = [cell.value for cell in workbook["Cars_Raw"][1]]
    assert "Financing" in raw_headers
    assert "Finanzierung" in raw_headers
    vehicle_headers = [cell.value for cell in workbook["Vehicles"][1]]
    assert "source_vehicle_url" in vehicle_headers
    assert "run_id" in vehicle_headers
    compliance_headers = [cell.value for cell in workbook["Requirements_Compliance"][1]]
    assert compliance_headers == [
        "field_name",
        "dataset",
        "required_by_task",
        "final_excel_column_exists",
        "extractor_exists",
        "current_source",
        "current_coverage_pct",
        "sample_present_value",
        "missing_count",
        "risk_level",
        "status",
        "notes",
    ]
    compliance_rows = {
        row[0].value: [cell.value for cell in row]
        for row in workbook["Requirements_Compliance"].iter_rows(min_row=2)
        if row[0].value
    }
    assert compliance_rows["Baureihe"][10] == "schema_only"
    assert compliance_rows["Baureihe"][9] == "high"


def test_word_report_accepts_phase1_summary_inputs(tmp_path):
    path = tmp_path / "report.docx"
    generate_word_report(
        path,
        pd.DataFrame([{"Händler ID": "C0000001"}]),
        pd.DataFrame([{"Vehicle_URL": "https://example.test"}]),
        {},
        "nordrhein-westfalen",
        [],
        run_summary={"run_id": "test-run", "vendors": 1, "vehicles_raw": 1, "errors": 0},
        vendor_coverage=pd.DataFrame(
            [{"field": "Händler ID", "non_empty_count": 1, "total_count": 1, "coverage_pct": 100.0}]
        ),
        vehicle_coverage=pd.DataFrame(
            [{"field": "Vehicle_URL", "non_empty_count": 1, "total_count": 1, "coverage_pct": 100.0}]
        ),
    )

    document = Document(path)
    text = "\n".join(paragraph.text for paragraph in document.paragraphs)
    assert "Data Completeness" in text
    assert "Schema completeness is guaranteed; source completeness is measured." in text
    assert "Manufacturer grouping follows the task-defined categories" in text
    assert "Unavailable source values are not guessed." in text
    assert "Some vehicle detail fields such as CO₂-Emissionen, Baureihe, Ausstattungslinie, and Anzahl der Fahrzeughalter showed low source coverage because mobile.de detail pages returned site-side 403/503 responses." in text
    assert "These values are not guessed and are reported transparently in Data_Coverage and Requirements_Compliance." in text
    assert "test-run" in text


def test_run_summary_distinguishes_discovered_enqueued_processed_and_detail_failures(tmp_path):
    config = ScraperConfig(project_root=tmp_path, max_vendors=5)
    config.regional_discovered_count = 32
    config.enqueued_vendor_count = 5
    config.cookie_modal_visible_count = 2
    config.cookie_consent_click_count = 1
    config.cookie_modal_remaining_count = 0
    config.vehicle_detail_jobs_total = 2
    config.detail_needed_count = 1
    config.detail_skipped_count = 1
    config.detail_attempted_count = 1
    config.detail_success_count = 0
    config.detail_failed_count = 1
    config.detail_site_blocked_or_503_count = 1
    config.playwright_browser_opened_count = 3
    config.regional_browser_opened_count = 1
    config.vendor_browser_opened_count = 1
    config.vehicle_detail_browser_opened_count = 1
    config.idle_about_blank_count = 0
    started = datetime(2026, 1, 1, tzinfo=timezone.utc)
    finished = started + timedelta(seconds=10)
    vendors = pd.DataFrame([{"Händler ID": f"C{i:07d}"} for i in range(5)])
    cars = pd.DataFrame(
        [
            {"Vehicle_URL": "a", "vehicle_data_source": "listing_fallback"},
            {"Vehicle_URL": "b", "vehicle_data_source": "detail_page"},
        ]
    )
    errors = [
        {
            "stage": "vehicle_detail_fetch_failed",
            "url": "https://example.test/1",
            "status_code": 503,
            "error_message": "fallback_reason=HTTP 403 final_status=HTTP 503",
        }
    ]

    summary = _compute_run_summary(
        "run-1",
        started,
        finished,
        config,
        vendors,
        cars,
        cars,
        errors,
    )

    assert summary["regional_discovered_count"] == 32
    assert summary["pipeline_mode"] == "sqlite"
    assert summary["strict_headless_blocked"] is False
    assert summary["enqueued_vendor_count"] == 5
    assert summary["processed_vendor_count"] == 5
    assert summary["processed_vendor_count"] <= config.max_vendors
    assert summary["detail_fetch_403_count"] == 1
    assert summary["detail_fetch_503_count"] == 1
    assert summary["detail_fetch_failed_count"] == 1
    assert summary["detail_site_blocked_or_503_count"] == 1
    assert summary["listing_fallback_used_count"] == 1
    assert "vendor_required_fields_coverage_pct" in summary
    assert "vehicle_basic_fields_coverage_pct" in summary
    assert "vehicle_technical_fields_coverage_pct" in summary
    assert "financing_fields_coverage_pct" in summary
    assert "classification_fields_coverage_pct" in summary
    assert "final_required_fields_coverage_pct" in summary
    assert summary["cookie_modal_visible_count"] == 2
    assert summary["cookie_consent_click_count"] == 1
    assert summary["cookie_modal_remaining_count"] == 0
    assert summary["vehicle_detail_jobs_total"] == 2
    assert summary["detail_needed_count"] == 1
    assert summary["detail_skipped_count"] == 1
    assert summary["detail_attempted_count"] == 1
    assert summary["detail_success_count"] == 0
    assert summary["detail_failed_count"] == 1
    assert summary["playwright_browser_opened_count"] == 3
    assert summary["regional_browser_opened_count"] == 1
    assert summary["vendor_browser_opened_count"] == 1
    assert summary["vehicle_detail_browser_opened_count"] == 1
    assert summary["idle_about_blank_count"] == 0
